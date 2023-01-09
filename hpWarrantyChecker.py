from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
import time
import random

# open the txt file and read in the serial numbers
with open('serial_numbers.txt', 'r') as f:
    serial_numbers = f.readlines()

# try to open the workbook
try:
    wb = openpyxl.load_workbook('serial_numbers_blocks.xlsx')
# if the workbook does not exist, create a new workbook
except FileNotFoundError:
    # create a new workbook
    wb = openpyxl.Workbook()
    # save the workbook with the specified name
    wb.save('serial_numbers_blocks.xlsx')

sheet = wb['Sheet']
# find the last row with data in it
last_row = sheet.max_row

# initialize a counter to keep track of which row we are on
row = last_row #+ 1
counter = 0
# loop through the serial numbers in blocks of 20
while counter <= len(serial_numbers):
    driver = webdriver.Firefox(service=Service(GeckoDriverManager().install()))
    driver.maximize_window()
    driver.get('https://support.hp.com/pt-pt/check-warranty#multiple')
    #time.sleep(3)
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="onetrust-accept-btn-handler"]')))
        time.sleep(random.uniform(0.5, 1.5))
    finally:
        cookie_btn = driver.find_element(By.XPATH, '//*[@id="onetrust-accept-btn-handler"]')
        cookie_btn.click()
        time.sleep(random.uniform(0.5, 1.5))
    
    for f in range(18):
        add_item_button = driver.find_element(By.ID, 'AddItem')
        add_item_button.click()
        time.sleep(random.uniform(0.5, 1.5))

# Iterate over the next 20 serial numbers
    for i in range(counter, counter+20):
        # Find the text box and enter the serial number
        if i == counter:
            text_box = driver.find_element(By.ID, 'inputtextpfinder')
        else:
            text_box = driver.find_element(By.ID, f'inputtextpfinder{i-counter}')
        text_box.send_keys(serial_numbers[i])
        time.sleep(random.uniform(0.5, 1.5))
    driver.find_element(By.ID, 'FindMyProduct').submit()
    counter +=20 
    
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.product-warranty-container')))
    data_divs = driver.find_elements(By.CSS_SELECTOR, 'div.product-warranty-container')
    for i, div in enumerate(data_divs):
        columns = div.text.split('\n')
        for j, column in enumerate(columns):
            sheet.cell(row=i+row, column=j+1).value = column
    print(f'Writing rows {row} to {row+19}')
    wb.save('serial_numbers_blocks.xlsx')
    row += 20
    time.sleep(random.uniform(0.5, 2))
    driver.quit()